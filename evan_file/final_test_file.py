from algorithm_version_1.constant_workspace_algorithm import ConstantWorkspaceAlgorithm
from algorithm_version_1.rectangle_bounding_box import RectangleBoundingBox
import csv
from openpyxl import load_workbook
import tracemalloc
import time
import gc

k = 7

while k < 8:
    if k == 0:
        n = 1100
        column_time = "AQ"
        column_memory = "AR"
    elif k == 1:
        n = 1200
        column_time = "AS"
        column_memory = "AT"
    elif k == 2:
        n = 1300
        column_time = "AU"
        column_memory = "AV"
    elif k == 3:
        n = 1400
        column_time = "AW"
        column_memory = "AX"
    elif k == 4:
        n = 1500
        column_time = "AY"
        column_memory = "AZ"
    elif k == 5:
        n = 2000
        column_time = "BA"
        column_memory = "BB"
    elif k == 6:
        n = 2500
        column_time = "BC"
        column_memory = "BD"
    elif k == 7:
        n = 3000
        column_time = "BE"
        column_memory = "BF"

    print(f"current n = {n}")

    seed = 18
    iteration_max = 18

    while seed < iteration_max + 1:
        print(f"current seed = {seed}")
        gc.disable()

        # Define a set of points
        points = []

        with open(f'INPUT/{n}points_seeds.csv', 'r', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if int(row['seed']) == seed:
                    x = float(row['x'])
                    y = float(row['y'])
                    points.append((x, y))

        # print(f"len(points): {len(points)}")

        # Define a bounding box as big as possible
        rectangleBoundingBox = RectangleBoundingBox(
            (float("-1000000"), float("1000000")),
            (float("-1000000"), float("1000000"))
        )

        tracemalloc.start()

        # Initialize the algorithm_version_2
        v = ConstantWorkspaceAlgorithm(rectangleBoundingBox)

        start_time = time.perf_counter()

        # Create the diagram
        flag = v.create_diagram(points=points, vis_steps=False, vis_result=False)
        if flag:
            print(f"seed = {seed} FAILED!")
            seed += 1
            continue
        else:
            end_time = time.perf_counter()

            total_memory_current, total_memory_peak = tracemalloc.get_traced_memory()

            del points
            del rectangleBoundingBox

            v.rectangleBoundingBox = None
            v.edges = None
            v.points = None

            gc.collect()
            after_del_memory_current, after_del_memory_peak = tracemalloc.get_traced_memory()

            tracemalloc.stop()
            gc.enable()

            read_only_and_write_only_memory = total_memory_current - after_del_memory_current

            read_write_memory = total_memory_peak - read_only_and_write_only_memory

            execution_time = end_time - start_time

            # print(f"Execution Time: {end_time - start_time:.6f} seconds (s)")
            # print(f"Current Pure Memory Usage: {pure_current / 1024:.4f} KB")

            if read_write_memory / 1024 > 5:
                print(f"Invalid Sample!")
                seed += 1
                continue
            else:
                excel_path = 'analysis.xlsx'
                wb = load_workbook(excel_path)
                ws = wb.active
                row = 3 + seed
                ws[f'{column_time}{row}'] = round(execution_time, 6)  # 秒
                ws[f'{column_memory}{row}'] = round(read_write_memory / 1024, 4)  # KB
                wb.save(excel_path)

                seed += 1
    k += 1
    print(f"### n = {n} Finished! ###")

print(f"ALL FINISHED!")